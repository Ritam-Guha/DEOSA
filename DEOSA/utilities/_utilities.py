# from DEOSA.fitness_functions import fitness_knapsack, fitness_fs
import numpy as np
from sklearn.neighbors import KNeighborsClassifier as KNN
import openpyxl as xl
from sklearn.model_selection import KFold

def get_transfer_function(shape):
    assert(shape in ["u", "s", "v"])
    if shape == "u":
        return Ufunction
    elif shape == "s":
        return sigmoid
    else:
        return Vfunction


def Ufunction(ip):
    # U-shaped transfer function
    alpha = 2
    beta = 1.5
    op = alpha * pow(abs(ip), beta)
    return op


def Vfunction(ip):
    # V-shaped transfer function
    op = 1 + ip*ip
    op = np.sqrt(op)
    op = ip/op
    return abs(op)


def sigmoid(ip):     
    # S-shaped transfer function
    if ip < 0:
        return 1 - 1/(1 + np.exp(ip))
    else:
        return 1/(1 + np.exp(-ip))


def sign_func(x): 
    # Signum function
    if x < 0:
        return -1
    return 1


def onecount(particle):
    # counts the number of features in a particle
    return int(np.sum(particle))


def compute_accuracy(data,
                     label,
                     particle,
                     seed=0):
    # function to compute classification accuracy  
    cols = np.flatnonzero(particle)     
    if cols.shape[0] == 0:
        return 0

    kf = KFold(n_splits=5, random_state=seed, shuffle=True)
    kf.get_n_splits(data)

    val = 0

    for train_index, test_index in kf.split(data):
        train_X, test_X = data[train_index], data[test_index]
        train_Y, test_Y = label[train_index], label[test_index]

        clf = KNN(n_neighbors=5)
        train_data = train_X[:, cols]
        test_data = test_X[:, cols]
        clf.fit(train_data, train_Y)
        val += clf.score(test_data, test_Y)

    val /= 5
    return val


def avg_concentration(eq_pool, pool_size, dimension):    
    # computes average concentration of the equilibrium pool
    avg = np.sum(eq_pool[0:pool_size, :], axis=0)
    avg = avg/pool_size
    return avg


def find_neighbor(particle, percent=0.3):   
    # forms neighbors of the given particle
    current_particle = particle.copy()    
    dimension = current_particle.shape[1]
    num_change = int(dimension*percent)
    pos = np.random.randint(0,dimension-1,num_change)
    current_particle[0, pos] = 1 - current_particle[0, pos]
    return current_particle    


def initialize(num_agents, num_features):
    # define min and max number of features
    min_features = int(0.3 * num_features)
    max_features = int(0.6 * num_features)

    # initialize the agents with zeros
    agents = np.zeros((num_agents, num_features))

    # select random features for each agent
    for agent_no in range(num_agents):

        # find random indices
        cur_count = np.random.randint(min_features, max_features)
        temp_vec = np.random.rand(1, num_features)
        temp_idx = np.argsort(temp_vec)[0][0:cur_count]

        # select the features with the ranom indices
        agents[agent_no][temp_idx] = 1   

    return agents

    
def plot_graph(final_result, conv_plot, dataset):
    # plottings 
    processes = ['DEO', 'DEOSA']   

    ################# plotting parameter-variation graphs ##################
    accuracy_var_pop_DEO = []
    accuracy_var_pop_DEOSA = []


    for particle_count in particle_count_testing:
        for max_iter in max_iter_testing:
            DEO_key = dataset + "_DEO" + "_pop_" + str(particle_count) + "_iter_" + str(max_iter) + "_runs_" + str(max_runs)
            DEOSA_key = dataset + "_DEOSA" + "_pop_" + str(particle_count) + "_iter_" + str(max_iter) + "_runs_" + str(max_runs)

            accuracy_var_pop_DEO.append(final_result[DEO_key + "avg_accuracy"])
            accuracy_var_pop_DEOSA.append(final_result[DEOSA_key + "avg_accuracy"])

    process_dict = {'DEO':accuracy_var_pop_DEO, 'DEOSA':accuracy_var_pop_DEOSA}
    output_file = dataset + "_parameter_variation" + ".jpg"
    plotter(dataset_name=dataset, x=particle_count_testing, x_label="Population size", y_dict=process_dict, y_objects=processes, y_label="Accuracy", title=dataset, storage_destination="Parameter Variation", file_name=output_file)        
    ###################################################################
    
    

    ################# plotting convergence graphs #####################
     
    x_range = np.arange(0, max_iter, 1)

    process_dict = {'DEO':conv_plot["DEO"], 'DEOSA':conv_plot["DEOSA"]}    
    output_file = dataset + "_convergence" + ".jpg"
    plotter(dataset_name=dataset, x=x_range, x_label="#Iteration", y_dict=process_dict, y_objects=processes, y_label="Fitness", title=dataset, storage_destination="Convergence", file_name=output_file)

    ###################################################################


def save_excel(particle_count_testing, max_iter_testing=[], conv_plot=[], final_result="", dataset="", save_type="parameter variation", init=0):
    # code to save the results in an excel file
    
    if(save_type == "parameter variation"):                
        
        if(init==1):     
            wb = xl.Workbook()
            ws = wb.active
            wb.remove(ws)
            for max_iter in max_iter_testing:            
                ws = wb.create_sheet("Iter_" + str(max_iter))
                ws.title = "Iter_" + str(max_iter)
                ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1) 
                ws.cell(1,1).value = "population size"

                cur_row = 4
                cur_col = 1

                for particle_count in particle_count_testing:
                    ws.cell(cur_row, cur_col).value = particle_count
                    cur_row += 1

                
        else:
            wb = xl.load_workbook("Results/parameter_variation.xlsx")
            for max_iter in max_iter_testing:    
                ws = wb["Iter_" + str(max_iter)]
                cur_row = 1
                cur_col = ws.max_column + 1

                # setting headers
                ws.merge_cells(start_row=cur_row, start_column=cur_col, end_row=cur_row, end_column=cur_col+3) 
                ws.merge_cells(start_row=cur_row+1, start_column=cur_col, end_row=cur_row+1, end_column=cur_col+1)
                ws.merge_cells(start_row=cur_row+1, start_column=cur_col+2, end_row=cur_row+1, end_column=cur_col+3)

                ws.cell(cur_row, cur_col).value = dataset
                ws.cell(cur_row+1, cur_col).value = "DEO"
                ws.cell(cur_row+1, cur_col+2).value = "DEOSA"
                ws.cell(cur_row+2, cur_col).value = "Accuracy"
                ws.cell(cur_row+2, cur_col+1).value = "#features"
                ws.cell(cur_row+2, cur_col+2).value = "Accuracy"
                ws.cell(cur_row+2, cur_col+3).value = "#features"

                cur_row += 3            
                for particle_count in particle_count_testing:

                    DEO_key = dataset + "_DEO" + "_pop_" + str(particle_count) + "_iter_" + str(max_iter)
                    DEOSA_key = dataset + "_DEOSA" + "_pop_" + str(particle_count) + "_iter_" + str(max_iter)

                    ws.cell(cur_row, cur_col).value = final_result[DEO_key+"avg_accuracy"]
                    ws.cell(cur_row, cur_col+1).value = final_result[DEO_key+"avg_num_features"]
                    ws.cell(cur_row, cur_col+2).value = final_result[DEOSA_key+"avg_accuracy"]
                    ws.cell(cur_row, cur_col+3).value = final_result[DEOSA_key+"avg_num_features"]

                    cur_row += 1
                    
        wb.save("Results/parameter_variation.xlsx")   
                    
    elif(save_type == "convergence"):                

        if(init==1):     
            wb = xl.Workbook()
            ws = wb.active
            wb.remove(ws)
            
            for max_iter in max_iter_testing:            
                ws = wb.create_sheet("Iter_" + str(max_iter))
                ws.title = "Iter_" + str(max_iter)

                cur_row = 1
                cur_col = 1
                ws.merge_cells(start_row=cur_row, start_column=cur_col, end_row=cur_row+1, end_column=cur_col)
                ws.cell(cur_row, cur_col).value = "#Iteration"

                cur_row += 2

                for count in range(1, max_iter+1):
                    ws.cell(cur_row, cur_col).value = count
                    cur_row += 1


        else:
            wb = xl.load_workbook("Results/convergence.xlsx")
            for max_iter in max_iter_testing:    
                ws = wb["Iter_" + str(max_iter)]
                cur_row = 1
                cur_col = ws.max_column + 1

                # setting headers
                ws.merge_cells(start_row=cur_row, start_column=cur_col, end_row=cur_row, end_column=cur_col+1)                

                ws.cell(cur_row, cur_col).value = dataset
                ws.cell(cur_row+1, cur_col).value = "DEO"
                ws.cell(cur_row+1, cur_col+1).value = "DEOSA"                

                cur_row += 2            
                for iteration_no in range(max_iter):

                    conv_DEO = conv_plot["DEO"]
                    conv_DEOSA = conv_plot["DEOSA"]

                    ws.cell(cur_row, cur_col).value = conv_DEO[iteration_no]
                    ws.cell(cur_row, cur_col+1).value = conv_DEOSA[iteration_no]                    

                    cur_row += 1

        wb.save("Results/convergence.xlsx")    
